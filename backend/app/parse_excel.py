"""解析产品知识库 Excel 文件，生成结构化 JSON 数据"""
import json
import sys
from pathlib import Path
from openpyxl import load_workbook


def parse_concept_sheet(ws) -> list[dict]:
    """解析概念解释 sheet"""
    concepts = []
    current_category = ""
    current_dimension = ""
    for row in ws.iter_rows(min_row=3, values_only=True):
        vals = [str(c).strip() if c else "" for c in row]
        if len(vals) < 5:
            continue
        _, category, dimension, classification, description = vals[:5]
        if category:
            current_category = category
        if dimension:
            current_dimension = dimension
        if classification and description:
            concepts.append({
                "category": current_category,
                "dimension": current_dimension,
                "classification": classification,
                "description": description,
            })
    return concepts


def parse_product_sheet(ws) -> list[dict]:
    """解析产品列表 sheet，动态读取表头"""
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 3:
        return []

    # 找到表头行（第2行，index=1）
    # 有些 sheet 第1行可能有合并的说明行，表头在第2行
    header_row = None
    for i, row in enumerate(rows):
        vals = [str(c).strip() if c else "" for c in row]
        if "序号" in vals:
            header_row = i
            break
    if header_row is None:
        return []

    headers = [str(c).strip() if c else f"col_{j}" for j, c in enumerate(rows[header_row])]

    products = []
    current_seq = ""
    current_offer = ""
    for row in rows[header_row + 1:]:
        vals = [str(c).strip() if c else "" for c in row]
        # 跳过全空行
        if all(v == "" or v == "None" for v in vals):
            continue

        product = {}
        for j, header in enumerate(headers):
            if j < len(vals) and vals[j] and vals[j] != "None":
                product[header] = vals[j]

        # 处理合并单元格：序号和 Offer 单号可能为空（子款）
        if "序号" in product and product["序号"]:
            current_seq = product["序号"]
            current_offer = product.get("Offer单号", current_offer)
        else:
            product["序号"] = current_seq
            if "Offer单号" not in product:
                product["Offer单号"] = current_offer

        # 移除空字段和图片列
        product.pop("图片", None)
        product.pop("", None)
        product = {k: v for k, v in product.items() if v and k}

        if "系统品名" in product:
            products.append(product)

    return products


def parse_excel(file_path: str) -> dict:
    """解析完整 Excel 文件"""
    wb = load_workbook(file_path)

    # 按品类配对 sheet
    category_map = {
        "床垫": {"concept_sheet": "概念解释（床垫）", "product_sheet": "床垫"},
        "智能马桶": {"concept_sheet": "概念解释（智能马桶）", "product_sheet": "智能马桶"},
        "花洒套装": {"concept_sheet": "概念解释（花洒套装）", "product_sheet": "花洒套装"},
        "智能门锁": {"concept_sheet": "概念解释（智能门锁）", "product_sheet": "智能门锁"},
        "吸顶灯": {"concept_sheet": "概念解释（吸顶灯）", "product_sheet": "吸顶灯"},
        "台灯": {"concept_sheet": "概念解释（台灯）", "product_sheet": "台灯"},
        "落地护眼灯": {"concept_sheet": "概念解释（落地护眼灯）", "product_sheet": "落地护眼灯"},
    }

    knowledge_base = {}
    for category, sheets in category_map.items():
        concepts = []
        products = []
        if sheets["concept_sheet"] in wb.sheetnames:
            concepts = parse_concept_sheet(wb[sheets["concept_sheet"]])
        if sheets["product_sheet"] in wb.sheetnames:
            products = parse_product_sheet(wb[sheets["product_sheet"]])
        knowledge_base[category] = {
            "concepts": concepts,
            "products": products,
        }
        print(f"  {category}: {len(concepts)} 概念, {len(products)} 产品")

    return knowledge_base


if __name__ == "__main__":
    excel_path = sys.argv[1] if len(sys.argv) > 1 else "/Users/hansonyang/Desktop/产品挑选指南（选品团队）.xlsx"
    print(f"正在解析: {excel_path}")
    kb = parse_excel(excel_path)

    output_path = Path(__file__).parent / "knowledge_base.json"
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(kb, f, ensure_ascii=False, indent=2)
    print(f"\n知识库已保存到: {output_path}")
    total_products = sum(len(v["products"]) for v in kb.values())
    total_concepts = sum(len(v["concepts"]) for v in kb.values())
    print(f"总计: {total_concepts} 个概念, {total_products} 个产品")
